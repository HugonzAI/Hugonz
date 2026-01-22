#!/usr/bin/env python3
"""
EST Converter Baseline V3.3

Converts Fluke ESA615 CSV test results to InterfaceTransactions XLSX format.

Architecture:
    1. Pure Rules Layer - Standard/ClassType inference, limits, canonicalization
    2. Parsing Layer - CSV ‚Üí FlukeParsed model
    3. Mapping Layer - FlukeParsed ‚Üí InterfaceRow
    4. Output Layer - Append to template XLSX
    5. UI Layer - PySide6 interface

Core Promises (DO NOT MODIFY without regression proof):
    - parse_fluke_file()
    - build_interface_row()
    - load_fixed_limits_from_summary() / get_fixed_limit_for_field()
    - append_rows_to_template()
"""

import sys
import os
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QListWidget, QLabel, QLineEdit, QFileDialog,
    QMessageBox, QListWidgetItem, QTabWidget
)
from PySide6.QtCore import Qt, Signal, QThread, QSettings
from PySide6.QtGui import QDragEnterEvent, QDropEvent, QIcon, QPixmap

import openpyxl
from openpyxl import load_workbook
from dateutil import parser as date_parser

# ESA615 Extension Module (optional - requires pyserial)
try:
    from esa615_ui_addon import ESA615Widget
    ESA615_AVAILABLE = True
except ImportError:
    ESA615_AVAILABLE = False
    print("Warning: ESA615 module not available (requires pyserial)")


# ============================================================================
# SECTION 1: DATA MODELS
# ============================================================================

@dataclass
class Measurement:
    """Single measurement from Fluke CSV."""
    group: str          # Canonical group (G_PROT_EARTH_RES, etc.)
    cond: str           # Normalized condition (NORMAL CONDITION, OPEN NEUTRAL, OPEN EARTH)
    subtest: str        # Original test name
    value: float        # Cleaned numeric value
    pf: str             # Pass/Fail status


@dataclass
class FlukeParsed:
    """Parsed Fluke CSV file data."""
    operator: str
    equipment: str
    tester_sn: str
    template: str
    dt: datetime
    measurements: List[Measurement] = field(default_factory=list)


@dataclass
class InterfaceRow:
    """Output row for InterfaceTransactions XLSX (22 columns)."""
    operator: str = ""
    equipment_number: str = ""
    asset_number: str = ""
    standard: str = ""
    test_type: str = ""
    class_type: str = ""
    test_seq: str = ""
    test_date: str = ""
    visual_pass_fail: str = ""
    line_load: str = "NA"
    earth_bond: str = "NA"
    insulation: str = "NA"
    earth_leakage_nc: str = "NA"
    earth_leakage_no: str = "NA"
    enclosure_leakage_nc: str = "NA"
    enclosure_leakage_no: str = "NA"
    enclosure_leakage_eo: str = "NA"
    applied_part_leakage_nc: str = "NA"
    applied_part_leakage_no: str = "NA"
    applied_part_leakage_eo: str = "NA"
    mains_contact: str = "NA"
    overall_pass_fail: str = "P"


# ============================================================================
# SECTION 2: PURE RULES LAYER
# ============================================================================

# Canonical group constants
G_PROT_EARTH_RES = "G_PROT_EARTH_RES"
G_MAINS_V_LN = "G_MAINS_V_LN"
G_MAINS_V_NE = "G_MAINS_V_NE"
G_INSULATION = "G_INSULATION"
G_EARTH_LEAK = "G_EARTH_LEAK"
G_ENC_LEAK = "G_ENC_LEAK"
G_PATIENT_LEAK = "G_PATIENT_LEAK"
G_MAINS_ON_AP = "G_MAINS_ON_AP"

# Normalized conditions
COND_NC = "NORMAL CONDITION"
COND_NO = "OPEN NEUTRAL"
COND_EO = "OPEN EARTH"


def canonical_group(raw_group: str) -> Optional[str]:
    """
    Normalize raw Fluke group name to canonical key.

    Returns canonical group constant or None if unrecognized.
    """
    upper = raw_group.upper()

    # Earth bond / protective earth resistance
    if "PROTECTIVE EARTH" in upper or "EARTH BOND" in upper or "EARTH CONTINUITY" in upper:
        return G_PROT_EARTH_RES

    # Mains voltage L-N
    if "MAINS" in upper and ("L-N" in upper or "LN" in upper) and "VOLTAGE" in upper:
        return G_MAINS_V_LN

    # Mains voltage N-E
    if "MAINS" in upper and ("N-E" in upper or "NE" in upper) and "VOLTAGE" in upper:
        return G_MAINS_V_NE

    # Insulation
    if "INSULATION" in upper:
        return G_INSULATION

    # Earth leakage
    if "EARTH" in upper and "LEAKAGE" in upper and "ENCLOSURE" not in upper:
        return G_EARTH_LEAK

    # Enclosure leakage
    if "ENCLOSURE" in upper and "LEAKAGE" in upper:
        return G_ENC_LEAK

    # Patient / applied part leakage
    if ("PATIENT" in upper or "APPLIED PART" in upper) and "LEAKAGE" in upper:
        return G_PATIENT_LEAK

    # Mains on applied parts
    if "MAINS" in upper and ("APPLIED" in upper or "PATIENT" in upper) and ("CONTACT" in upper or "ON" in upper):
        return G_MAINS_ON_AP

    return None


def normalize_condition(raw_cond: str) -> str:
    """
    Normalize condition string.

    Returns one of: NORMAL CONDITION, OPEN NEUTRAL, OPEN EARTH
    """
    upper = raw_cond.upper()

    if "OPEN NEUTRAL" in upper or "OPEN N" in upper or "O/N" in upper:
        return COND_NO

    if "OPEN EARTH" in upper or "OPEN E" in upper or "O/E" in upper:
        return COND_EO

    return COND_NC


def infer_standard(template_name: str) -> Optional[str]:
    """
    Infer AS/NZS standard from template name.

    Rules (in order):
        1. Contains "3760" ‚Üí AS/NZS 3760
        2. Contains "3551" ‚Üí AS/NZS 3551
        3. "NO EARTH" + "DOMESTIC" ‚Üí AS/NZS 3760
        4. "NO EARTH" (non-domestic) ‚Üí AS/NZS 3551
        5. Token "1D/2D/5D" ‚Üí AS/NZS 3760
        6. Token "5B/5BF/5CF/5C" or "TYPE BF/CF/B" ‚Üí AS/NZS 3551
        7. Else ‚Üí None (inference failed)

    Returns: "AS/NZS 3551" or "AS/NZS 3760" or None
    """
    upper = template_name.upper()
    tokens = re.findall(r'\b\w+\b', upper)

    # Rule 1: Explicit 3760
    if "3760" in upper:
        return "AS/NZS 3760"

    # Rule 2: Explicit 3551
    if "3551" in upper:
        return "AS/NZS 3551"

    # Rule 3: NO EARTH + DOMESTIC
    if "NO EARTH" in upper and "DOMESTIC" in upper:
        return "AS/NZS 3760"

    # Rule 4: NO EARTH (non-domestic)
    if "NO EARTH" in upper:
        return "AS/NZS 3551"

    # Rule 5: Tokens indicating 3760
    for token in tokens:
        if token in ["1D", "2D", "5D"]:
            return "AS/NZS 3760"

    # Rule 6: Tokens indicating 3551
    for token in tokens:
        if token in ["5B", "5BF", "5CF", "5C"] or "TYPE" in upper and token in ["BF", "CF", "B"]:
            return "AS/NZS 3551"

    return None


def infer_class_type(template_name: str, standard: str) -> str:
    """
    Infer ClassType from template name and standard.

    Args:
        template_name: Original template name from CSV
        standard: Inferred standard (AS/NZS 3551 or AS/NZS 3760)

    Returns: ClassType string (e.g., "1D", "5BF", "5CF&5BF", etc.)
    """
    upper = template_name.upper()
    tokens = re.findall(r'\b\w+\b', upper)

    # NO EARTH cases
    if "NO EARTH" in upper:
        # NO EARTH DOMESTIC ‚Üí 5D
        if "DOMESTIC" in upper:
            return "5D"

        # NO EARTH + (5CF & (5BF or BF)) ‚Üí 5CF&5BF
        has_5cf = "5CF" in tokens
        has_bf = "5BF" in tokens or "BF" in tokens
        if has_5cf and has_bf:
            return "5CF&5BF"

        # NO EARTH + NO AP ‚Üí 5
        if "NO AP" in upper or "NO APPLIED" in upper:
            return "5"

        # NO EARTH + TYPE CF ‚Üí 5CF
        if "TYPE CF" in upper or "CF" in tokens:
            return "5CF"

        # NO EARTH + TYPE BF ‚Üí 5BF
        if "TYPE BF" in upper or "BF" in tokens:
            return "5BF"

        # NO EARTH + TYPE B ‚Üí 5B
        if "TYPE B" in upper or "B" in tokens:
            return "5B"

        # Default NO EARTH
        return "5"

    # Earthed cases
    # Extract base class number (1, 2, 5, etc.)
    base_class = None
    for token in tokens:
        if token.isdigit() and token in ["1", "2", "5"]:
            base_class = token
            break
        # Check for combined tokens like "1D", "2D"
        if re.match(r'^[125][A-Z]+$', token):
            base_class = token[0]
            break

    if base_class is None:
        base_class = "1"  # Default

    if standard == "AS/NZS 3760":
        return f"{base_class}D"

    # 3551 earthed - check for type
    if "TYPE CF" in upper or "CF" in tokens:
        return f"{base_class}CF"
    if "TYPE BF" in upper or "BF" in tokens:
        return f"{base_class}BF"
    if "TYPE B" in upper:
        return f"{base_class}B"

    # Default 3551 earthed
    return f"{base_class}BF"


def class_is_earthed(class_type: str, standard: str) -> bool:
    """
    Determine if class type requires earth bond.

    Args:
        class_type: Class type string
        standard: AS/NZS standard

    Returns: True if earthed class, False otherwise
    """
    # NO EARTH classes
    if class_type.startswith("5") and standard == "AS/NZS 3551":
        return False

    return True


# Default fixed limits (fallback if EST_Limits_Summary.xlsx not found)
DEFAULT_LIMITS_3551 = {
    "earth_bond": "0.2",
    "insulation": "1.0",
    "earth_leakage_nc": "5000",
    "earth_leakage_no": "5000",
    "enclosure_leakage_nc": "500",
    "enclosure_leakage_no": "500",
    "enclosure_leakage_eo": "500",
    "patient_leakage_nc": "500",
    "patient_leakage_no": "500",
    "patient_leakage_eo": "500",
    "mains_contact": "25000",
}

DEFAULT_LIMITS_3760 = {
    "earth_bond": "1",
    "insulation": "1.0",
    "earth_leakage_nc": "5000",
    "earth_leakage_no": "5000",
    "enclosure_leakage_nc": "1000",  # Class 2
    "enclosure_leakage_no": "1000",
    "enclosure_leakage_eo": "1000",
    # Patient leakage / mains on AP ‚Üí NA for 3760 baseline
}


def get_default_limits(standard: str, class_type: str) -> Dict[str, str]:
    """
    Get default limits for given standard and class type.

    Args:
        standard: AS/NZS standard
        class_type: Class type

    Returns: Dictionary of field ‚Üí limit value
    """
    if standard == "AS/NZS 3760":
        return DEFAULT_LIMITS_3760.copy()
    else:
        return DEFAULT_LIMITS_3551.copy()


# Global limits cache (loaded from EST_Limits_Summary.xlsx)
_LIMITS_CACHE: Dict[str, Dict[str, str]] = {}


def load_fixed_limits_from_summary(summary_path: str = "EST_Limits_Summary.xlsx"):
    """
    Load fixed limits from EST_Limits_Summary.xlsx.

    Expected structure:
        Sheet: Summary or first sheet
        Columns: Standard, ClassType, Field, Limit

    Populates global _LIMITS_CACHE.
    """
    global _LIMITS_CACHE
    _LIMITS_CACHE = {}

    if not os.path.exists(summary_path):
        print(f"Warning: {summary_path} not found, using default limits")
        return

    try:
        wb = load_workbook(summary_path, data_only=True)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
        else:
            ws = wb.active

        # Find header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            if row and "Standard" in str(row) and "Field" in str(row):
                header_row = row_idx
                break

        if header_row is None:
            print("Warning: Could not find header in limits summary")
            return

        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue

            standard = str(row[0]).strip()
            class_type = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            field = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
            limit = str(row[3]).strip() if len(row) > 3 and row[3] else ""

            if not standard or not field or not limit:
                continue

            key = f"{standard}|{class_type}|{field}"
            _LIMITS_CACHE[key] = limit

        print(f"Loaded {len(_LIMITS_CACHE)} limits from {summary_path}")

    except Exception as e:
        print(f"Error loading limits summary: {e}")


def get_fixed_limit_for_field(standard: str, class_type: str, field: str) -> str:
    """
    Get fixed limit for a specific field.

    Args:
        standard: AS/NZS standard
        class_type: Class type
        field: Field name (e.g., "earth_bond", "insulation")

    Returns: Limit value as string
    """
    # Try exact match
    key = f"{standard}|{class_type}|{field}"
    if key in _LIMITS_CACHE:
        return _LIMITS_CACHE[key]

    # Try without class type (generic for standard)
    key = f"{standard}||{field}"
    if key in _LIMITS_CACHE:
        return _LIMITS_CACHE[key]

    # Fallback to defaults
    defaults = get_default_limits(standard, class_type)
    return defaults.get(field, "")


# ============================================================================
# SECTION 3: PARSING LAYER
# ============================================================================

def clean_value(value_str: str) -> Optional[float]:
    """
    Extract numeric value from Fluke value string.

    Examples:
        "0.15 Ohm" ‚Üí 0.15
        "1.2" ‚Üí 1.2
        "> 1000 MOhm" ‚Üí 1000.0

    Returns: Float value or None if parsing fails
    """
    if not value_str:
        return None

    # Remove common units and symbols
    cleaned = re.sub(r'[A-Za-z\s><=]', '', value_str)
    cleaned = cleaned.strip()

    try:
        return float(cleaned)
    except ValueError:
        return None


def _find_results_column_map(header_row: List[str]) -> Tuple[int, int]:
    """
    Find value and status column indices in results header.

    Args:
        header_row: List of header cell values

    Returns: Tuple of (value_col_idx, status_col_idx)
    """
    value_col = 5  # Default fallback
    status_col = 8  # Default fallback

    for idx, cell in enumerate(header_row):
        if not cell:
            continue
        cell_upper = str(cell).upper()
        if "VALUE" in cell_upper or "RESULT" in cell_upper:
            value_col = idx
        if "STATUS" in cell_upper or "PASS" in cell_upper or "P/F" in cell_upper:
            status_col = idx

    return value_col, status_col


def parse_fluke_file(csv_path: str) -> Tuple[Optional[FlukeParsed], Optional[str]]:
    """
    Parse Fluke ESA615 CSV file.

    Args:
        csv_path: Path to CSV file

    Returns:
        Tuple of (FlukeParsed object, error message)
        If successful: (FlukeParsed, None)
        If failed: (None, error_message)
    """
    try:
        # Try multiple encodings to handle different CSV file formats
        rows = None
        last_error = None
        for encoding in ['utf-8-sig', 'utf-8', 'windows-1252', 'latin-1', 'iso-8859-1']:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                break  # Success, exit loop
            except (UnicodeDecodeError, UnicodeError) as e:
                last_error = e
                continue  # Try next encoding

        if rows is None:
            return None, f"Could not decode CSV file with any common encoding. Last error: {last_error}"

        if len(rows) < 10:
            return None, "CSV file too short"

        # Extract metadata (first ~20 rows)
        operator = ""
        equipment = ""
        tester_sn = ""
        template = ""
        test_date = ""

        for row in rows[:20]:
            if len(row) < 2:
                continue

            # Check all columns in the row for key-value pairs
            # Fluke CSV format may have multiple key-value pairs per row
            for i in range(len(row)):
                if not row[i] or ':' not in str(row[i]):
                    continue

                key = str(row[i]).strip().upper()
                # Value is typically 2 columns after key (skip empty column)
                value_idx = i + 2 if i + 2 < len(row) else i + 1
                value = str(row[value_idx]).strip() if value_idx < len(row) and row[value_idx] else ""

                if not value:  # If i+2 is empty, try i+1
                    value_idx = i + 1
                    value = str(row[value_idx]).strip() if value_idx < len(row) and row[value_idx] else ""

                if "OPERATOR" in key and not operator:
                    operator = value
                elif ("EQUIPMENT" in key or "ASSET" in key) and not equipment:
                    equipment = value
                elif "SERIAL" in key and i < 5 and not tester_sn:  # Serial Number in left columns = Tester S/N
                    tester_sn = value
                elif "TEMPLATE" in key and not template:
                    template = value
                elif "DATE" in key and "TIME" in key and not test_date:
                    test_date = value

        # Validate required fields
        if not equipment:
            return None, "Missing Equipment Number"
        if not tester_sn:
            return None, "Missing Tester S/N"
        if not template:
            return None, "Missing Template Name"

        # Parse datetime
        dt = None
        if test_date:
            try:
                dt = date_parser.parse(test_date)
            except:
                dt = datetime.now()
        else:
            dt = datetime.now()

        # Find results header
        results_start_row = None
        for idx, row in enumerate(rows):
            if len(row) > 0 and str(row[0]).strip().upper() == "TEST NAME":
                # Check if this row has Value column
                if any("VALUE" in str(cell).upper() or "RESULT" in str(cell).upper() for cell in row):
                    results_start_row = idx
                    break

        if results_start_row is None:
            return None, "Could not find results header"

        header_row = rows[results_start_row]
        value_col, status_col = _find_results_column_map(header_row)

        # Parse measurements
        measurements = []
        for row in rows[results_start_row + 1:]:
            if len(row) <= max(value_col, status_col):
                continue

            test_name = str(row[0]).strip() if row[0] else ""
            if not test_name:
                continue

            # Extract group and condition from test name
            # Format: "Group - Condition - Subtest" or variations
            parts = [p.strip() for p in test_name.split('-')]

            raw_group = parts[0] if len(parts) > 0 else ""
            raw_cond = parts[1] if len(parts) > 1 else ""
            subtest = parts[2] if len(parts) > 2 else test_name

            # Canonicalize
            group = canonical_group(raw_group)
            if group is None:
                continue  # Skip unrecognized groups

            cond = normalize_condition(raw_cond)

            # Extract value and status
            value_str = str(row[value_col]).strip() if len(row) > value_col and row[value_col] else ""
            status_str = str(row[status_col]).strip().upper() if len(row) > status_col and row[status_col] else ""

            value = clean_value(value_str)
            if value is None:
                continue

            pf = "Pass" if status_str.startswith("P") else "Fail"

            measurements.append(Measurement(
                group=group,
                cond=cond,
                subtest=subtest,
                value=value,
                pf=pf
            ))

        parsed = FlukeParsed(
            operator=operator,
            equipment=equipment,
            tester_sn=tester_sn,
            template=template,
            dt=dt,
            measurements=measurements
        )

        return parsed, None

    except Exception as e:
        return None, f"Parse error: {str(e)}"


# ============================================================================
# SECTION 4: MAPPING LAYER
# ============================================================================

def build_interface_row(parsed: FlukeParsed, tester_map: Dict[str, str]) -> Tuple[Optional[InterfaceRow], List[str]]:
    """
    Build InterfaceRow from FlukeParsed data.

    Args:
        parsed: FlukeParsed object
        tester_map: Dictionary mapping Tester S/N ‚Üí Asset Number

    Returns:
        Tuple of (InterfaceRow object, list of errors)
        If successful: (InterfaceRow, [])
        If failed: (None, [error_messages])
    """
    errors = []

    # Infer standard
    standard = infer_standard(parsed.template)
    if standard is None:
        errors.append(f"Could not infer standard from template: {parsed.template}")
        return None, errors

    # Infer class type
    class_type = infer_class_type(parsed.template, standard)

    # Asset number mapping
    asset_number = tester_map.get(parsed.tester_sn, parsed.tester_sn)

    # Test sequence (strip prefix like "_0001 - ")
    test_seq = parsed.template
    test_seq = re.sub(r'^_\d+\s*-\s*', '', test_seq)

    # Format test date
    test_date = parsed.dt.strftime("%d/%m/%Y")

    # Initialize row
    row = InterfaceRow(
        operator=parsed.operator,
        equipment_number=parsed.equipment,
        asset_number=asset_number,
        standard=standard,
        test_type="Routine",
        class_type=class_type,
        test_seq=test_seq,
        test_date=test_date,
        visual_pass_fail="P"  # Assumed
    )

    # Determine if class is earthed
    is_earthed = class_is_earthed(class_type, standard)

    # Build measurement lookup
    meas_map = {}
    for m in parsed.measurements:
        key = (m.group, m.cond)
        if key not in meas_map:
            meas_map[key] = []
        meas_map[key].append(m)

    # Helper: Format field value
    def format_field(meas: Optional[Measurement], limit: str, unit: str) -> str:
        if meas is None:
            return "NA"
        pf_str = "Pass" if meas.pf == "Pass" else "Failed"
        return f"{meas.value},{pf_str},{limit},{unit}"

    # Line Load (requires both LN and NE)
    ln_meas = meas_map.get((G_MAINS_V_LN, COND_NC))
    ne_meas = meas_map.get((G_MAINS_V_NE, COND_NC))
    if ln_meas and ne_meas:
        ln_val = ln_meas[0].value
        ne_val = ne_meas[0].value
        row.line_load = f"{ln_val}V [L-N={ne_val}V][Load=0.0kVA]"

    # Earth Bond
    if is_earthed:
        eb_meas = meas_map.get((G_PROT_EARTH_RES, COND_NC))
        if eb_meas:
            limit = get_fixed_limit_for_field(standard, class_type, "earth_bond")
            row.earth_bond = format_field(eb_meas[0], limit, "Ohms")
    else:
        row.earth_bond = "NA"

    # Insulation
    ins_meas = meas_map.get((G_INSULATION, COND_NC))
    if ins_meas:
        limit = get_fixed_limit_for_field(standard, class_type, "insulation")
        # Extract voltage from subtest if present
        voltage_label = ""
        if ins_meas[0].subtest:
            if "250" in ins_meas[0].subtest:
                voltage_label = " [250V]"
            elif "500" in ins_meas[0].subtest:
                voltage_label = " [500V]"
        unit = f"MOhms{voltage_label}"
        row.insulation = format_field(ins_meas[0], limit, unit)

    # Earth Leakage
    if standard == "AS/NZS 3760" or (standard == "AS/NZS 3551" and is_earthed):
        el_nc = meas_map.get((G_EARTH_LEAK, COND_NC))
        el_no = meas_map.get((G_EARTH_LEAK, COND_NO))

        if el_nc:
            limit = get_fixed_limit_for_field(standard, class_type, "earth_leakage_nc")
            row.earth_leakage_nc = format_field(el_nc[0], limit, "uA")

        if el_no:
            limit = get_fixed_limit_for_field(standard, class_type, "earth_leakage_no")
            row.earth_leakage_no = format_field(el_no[0], limit, "uA")

    # Enclosure Leakage
    enc_nc = meas_map.get((G_ENC_LEAK, COND_NC))
    enc_no = meas_map.get((G_ENC_LEAK, COND_NO))
    enc_eo = meas_map.get((G_ENC_LEAK, COND_EO))

    if enc_nc:
        limit = get_fixed_limit_for_field(standard, class_type, "enclosure_leakage_nc")
        row.enclosure_leakage_nc = format_field(enc_nc[0], limit, "uA")

    if enc_no:
        limit = get_fixed_limit_for_field(standard, class_type, "enclosure_leakage_no")
        row.enclosure_leakage_no = format_field(enc_no[0], limit, "uA")

    if enc_eo:
        limit = get_fixed_limit_for_field(standard, class_type, "enclosure_leakage_eo")
        row.enclosure_leakage_eo = format_field(enc_eo[0], limit, "uA")

    # Patient / Applied Part Leakage (3551 only, baseline simplification for 3760)
    if standard == "AS/NZS 3551":
        pat_nc = meas_map.get((G_PATIENT_LEAK, COND_NC))
        pat_no = meas_map.get((G_PATIENT_LEAK, COND_NO))
        pat_eo = meas_map.get((G_PATIENT_LEAK, COND_EO))

        if pat_nc:
            limit = get_fixed_limit_for_field(standard, class_type, "patient_leakage_nc")
            row.applied_part_leakage_nc = format_field(pat_nc[0], limit, "uA")

        if pat_no:
            limit = get_fixed_limit_for_field(standard, class_type, "patient_leakage_no")
            row.applied_part_leakage_no = format_field(pat_no[0], limit, "uA")

        if pat_eo:
            limit = get_fixed_limit_for_field(standard, class_type, "patient_leakage_eo")
            row.applied_part_leakage_eo = format_field(pat_eo[0], limit, "uA")

    # Mains Contact (3551 only)
    if standard == "AS/NZS 3551":
        mc_meas = meas_map.get((G_MAINS_ON_AP, COND_NC))
        if mc_meas:
            limit = get_fixed_limit_for_field(standard, class_type, "mains_contact")
            row.mains_contact = format_field(mc_meas[0], limit, "uA")

    # Overall Pass/Fail
    for m in parsed.measurements:
        if m.pf == "Fail":
            row.overall_pass_fail = "F"
            break

    return row, errors


# ============================================================================
# SECTION 5: OUTPUT LAYER
# ============================================================================

def load_tester_map(tester_path: str = "EST Tester.xlsx") -> Dict[str, str]:
    """
    Load tester S/N ‚Üí Asset Number mapping from EST Tester.xlsx.

    Expected structure:
        Columns: Serial Number (or S/N), Asset Number

    Returns: Dictionary of S/N ‚Üí Asset Number
    """
    tester_map = {}

    if not os.path.exists(tester_path):
        print(f"Warning: {tester_path} not found, using S/N as Asset Number")
        return tester_map

    try:
        wb = load_workbook(tester_path, data_only=True)
        ws = wb.active

        # Find header row
        sn_col = None
        asset_col = None
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            if row:
                for col_idx, cell in enumerate(row):
                    if cell:
                        cell_upper = str(cell).upper()
                        if "SERIAL" in cell_upper or "S/N" in cell_upper:
                            sn_col = col_idx
                            header_row = row_idx
                        if "ASSET" in cell_upper:
                            asset_col = col_idx

        if sn_col is None or asset_col is None or header_row is None:
            print("Warning: Could not find S/N or Asset columns in tester file")
            return tester_map

        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) <= max(sn_col, asset_col):
                continue

            sn = str(row[sn_col]).strip() if row[sn_col] else ""
            asset = str(row[asset_col]).strip() if row[asset_col] else ""

            if sn and asset:
                tester_map[sn] = asset

        print(f"Loaded {len(tester_map)} tester mappings")

    except Exception as e:
        print(f"Error loading tester map: {e}")

    return tester_map


def append_rows_to_template(
    template_path: str,
    output_path: str,
    rows: List[InterfaceRow]
) -> Tuple[bool, Optional[str]]:
    """
    Append InterfaceRows to template XLSX and save as new file.

    Args:
        template_path: Path to example_Good.xlsx template
        output_path: Path to save output file
        rows: List of InterfaceRow objects to append

    Returns:
        Tuple of (success: bool, error_message: Optional[str])
    """
    try:
        # Load template
        wb = load_workbook(template_path)
        ws = wb.active

        # Find last row with data
        last_row = ws.max_row

        # Append rows
        for row in rows:
            last_row += 1
            ws.cell(last_row, 1, row.operator)
            ws.cell(last_row, 2, row.equipment_number)
            ws.cell(last_row, 3, row.asset_number)
            ws.cell(last_row, 4, row.standard)
            ws.cell(last_row, 5, row.test_type)
            ws.cell(last_row, 6, row.class_type)
            ws.cell(last_row, 7, row.test_seq)
            ws.cell(last_row, 8, row.test_date)
            ws.cell(last_row, 9, row.visual_pass_fail)
            ws.cell(last_row, 10, row.line_load)
            ws.cell(last_row, 11, row.earth_bond)
            ws.cell(last_row, 12, row.insulation)
            ws.cell(last_row, 13, row.earth_leakage_nc)
            ws.cell(last_row, 14, row.earth_leakage_no)
            ws.cell(last_row, 15, row.enclosure_leakage_nc)
            ws.cell(last_row, 16, row.enclosure_leakage_no)
            ws.cell(last_row, 17, row.enclosure_leakage_eo)
            ws.cell(last_row, 18, row.applied_part_leakage_nc)
            ws.cell(last_row, 19, row.applied_part_leakage_no)
            ws.cell(last_row, 20, row.applied_part_leakage_eo)
            ws.cell(last_row, 21, row.mains_contact)
            ws.cell(last_row, 22, row.overall_pass_fail)

        # Save
        wb.save(output_path)
        return True, None

    except Exception as e:
        return False, f"Output error: {str(e)}"


# ============================================================================
# SECTION 6: CONVERSION ORCHESTRATION
# ============================================================================

def convert_files(
    csv_paths: List[str],
    output_folder: str,
    template_path: str,
    progress_callback=None
) -> Tuple[int, int, List[Tuple[str, str]]]:
    """
    Convert multiple Fluke CSV files to single XLSX output.

    Args:
        csv_paths: List of CSV file paths
        output_folder: Output directory
        template_path: Path to example_Good.xlsx
        progress_callback: Optional callback(current, total, message)

    Returns:
        Tuple of (success_count, error_count, error_list)
        error_list: List of (filename, error_message) tuples
    """
    # Load dependencies
    load_fixed_limits_from_summary()
    tester_map = load_tester_map()

    # Results
    interface_rows = []
    errors = []

    total = len(csv_paths)

    for idx, csv_path in enumerate(csv_paths):
        filename = os.path.basename(csv_path)

        if progress_callback:
            progress_callback(idx + 1, total, f"Processing {filename}")

        # Parse CSV
        parsed, error = parse_fluke_file(csv_path)
        if parsed is None:
            errors.append((filename, error))
            continue

        # Build interface row
        row, row_errors = build_interface_row(parsed, tester_map)
        if row is None:
            error_msg = "; ".join(row_errors) if row_errors else "Unknown error"
            errors.append((filename, error_msg))
            continue

        interface_rows.append(row)

    # Generate output files
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Write successful rows to XLSX
    if interface_rows:
        output_xlsx = os.path.join(output_folder, f"EST_UPLOAD_{timestamp}.xlsx")
        success, error = append_rows_to_template(template_path, output_xlsx, interface_rows)
        if not success:
            errors.append(("Output XLSX", error))

    # Write error CSV if any errors
    if errors:
        error_csv = os.path.join(output_folder, f"EST_ERRORS_{timestamp}.csv")
        try:
            with open(error_csv, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["File", "Error"])
                writer.writerows(errors)
        except Exception as e:
            print(f"Failed to write error CSV: {e}")

    success_count = len(interface_rows)
    error_count = len(errors)

    return success_count, error_count, errors


# ============================================================================
# SECTION 7: UI LAYER
# ============================================================================

class ConvertThread(QThread):
    """Background thread for conversion to keep UI responsive."""

    progress = Signal(int, int, str)  # current, total, message
    finished = Signal(int, int, list)  # success_count, error_count, errors

    def __init__(self, csv_paths, output_folder, template_path):
        super().__init__()
        self.csv_paths = csv_paths
        self.output_folder = output_folder
        self.template_path = template_path

    def run(self):
        success, errors_count, errors = convert_files(
            self.csv_paths,
            self.output_folder,
            self.template_path,
            self.progress.emit
        )
        self.finished.emit(success, errors_count, errors)


class FileListWidget(QListWidget):
    """Custom list widget with drag-and-drop support and visual feedback."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragDropMode(QListWidget.InternalMove)
        self.drag_active = False

        # Set placeholder hint
        self.setMinimumHeight(200)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.drag_active = True
            self.setStyleSheet("""
                QListWidget {
                    background: #ecfdf5;
                    border: 3px dashed #10b981;
                    border-radius: 8px;
                    padding: 8px;
                }
                QListWidget::item {
                    padding: 12px;
                    border-radius: 6px;
                    margin: 3px 0;
                    background: white;
                    border: 1px solid #10b981;
                }
            """)

    def dragLeaveEvent(self, event):
        self.drag_active = False
        self.setStyleSheet("""
            QListWidget {
                background: #f9fafb;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                padding: 8px;
                font-size: 10pt;
            }
            QListWidget::item {
                padding: 12px;
                border-radius: 6px;
                margin: 3px 0;
                background: white;
                border: 1px solid #e5e7eb;
            }
            QListWidget::item:hover {
                background: #f0fdf4;
                border-color: #10b981;
            }
            QListWidget::item:selected {
                background: #d1fae5;
                border-color: #10b981;
                color: #065f46;
            }
        """)

    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                path = url.toLocalFile()
                if path.lower().endswith('.csv'):
                    self.add_file(path)
            event.acceptProposedAction()

        self.drag_active = False
        # Restore normal style
        self.dragLeaveEvent(None)

    def add_file(self, path: str):
        """Add file to list with filename display and full path tooltip."""
        filename = os.path.basename(path)

        # Check if file already exists
        for i in range(self.count()):
            if self.item(i).data(Qt.UserRole) == path:
                return  # Don't add duplicates

        item = QListWidgetItem(f"üìÑ {filename}")
        item.setToolTip(f"Full path: {path}\nClick to select")
        item.setData(Qt.UserRole, path)  # Store full path
        self.addItem(item)


class MainWindow(QMainWindow):
    """Main application window."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("EST Converter V3.3 - Electrical Safety Testing")
        self.setMinimumSize(1000, 700)

        # Try to load icon
        if os.path.exists("est_icon.ico"):
            self.setWindowIcon(QIcon("est_icon.ico"))

        # Apply modern stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background: #f9fafb;
            }
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 10pt;
            }
            QLabel {
                color: #111827;
            }
            QLineEdit {
                padding: 10px 15px;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                background: white;
                color: #111827;
                font-size: 10pt;
            }
            QLineEdit:focus {
                border: 2px solid #10b981;
                outline: none;
            }
            QPushButton {
                padding: 10px 20px;
                border: none;
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #10b981, stop:1 #059669);
                color: white;
                font-weight: 600;
                font-size: 10pt;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #059669, stop:1 #047857);
            }
            QPushButton:pressed {
                background: #047857;
            }
            QPushButton:disabled {
                background: #d1d5db;
                color: #9ca3af;
            }
        """)

        # Load settings
        self.settings = QSettings()

        self.init_ui()

        # Load last output folder from settings
        last_output = self.settings.value("output_folder", "")
        if last_output and os.path.exists(last_output):
            self.output_edit.setText(last_output)
            # Update ESA615Widget output directory if available
            if ESA615_AVAILABLE and hasattr(self, 'esa615_widget'):
                self.esa615_widget.output_dir = last_output

        # Fixed template path (no longer user-selectable)
        self.template_path = "example_Good.xlsx"

    def init_ui(self):
        """Initialize UI components."""
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header Section (Logo + Title in card)
        header_card = QWidget()
        header_card.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #10b981, stop:1 #3b82f6);
                border-radius: 15px;
                padding: 20px;
            }
        """)
        header_layout = QHBoxLayout(header_card)

        # Logo (optional)
        if os.path.exists("hnz_logo.png"):
            logo_label = QLabel()
            logo_label.setStyleSheet("background: transparent;")
            pixmap = QPixmap("hnz_logo.png")
            scaled = pixmap.scaledToHeight(60, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled)
            logo_label.setFixedSize(60, 60)
            logo_label.setScaledContents(False)
            header_layout.addWidget(logo_label)

        # Title with subtitle
        title_container = QWidget()
        title_container.setStyleSheet("background: transparent;")
        title_layout = QVBoxLayout(title_container)
        title_layout.setSpacing(5)
        title_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("EST Converter")
        title.setStyleSheet("""
            background: transparent;
            color: white;
            font-size: 24pt;
            font-weight: 700;
            letter-spacing: -1px;
        """)

        subtitle = QLabel("Fluke ESA615 Electrical Safety Testing")
        subtitle.setStyleSheet("""
            background: transparent;
            color: rgba(255, 255, 255, 0.9);
            font-size: 11pt;
            font-weight: 400;
        """)

        version_label = QLabel("Version 3.3 Baseline")
        version_label.setStyleSheet("""
            background: transparent;
            color: rgba(255, 255, 255, 0.7);
            font-size: 9pt;
            margin-top: 3px;
        """)

        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        title_layout.addWidget(version_label)
        header_layout.addWidget(title_container)
        header_layout.addStretch()

        layout.addWidget(header_card)

        # Tab widget for ESA615 and CSV files (shared display area)
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                background: white;
                border: 2px solid #e5e7eb;
                border-radius: 12px;
                padding: 15px;
                margin-top: -1px;
            }
            QTabBar::tab {
                background: #f3f4f6;
                color: #4b5563;
                padding: 12px 28px;
                margin-right: 4px;
                border: 2px solid #e5e7eb;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 600;
                font-size: 11pt;
                min-width: 120px;
            }
            QTabBar::tab:selected {
                background: white;
                color: #10b981;
                border-color: #e5e7eb;
                border-bottom: 2px solid white;
                margin-bottom: -2px;
            }
            QTabBar::tab:hover:!selected {
                background: #e5e7eb;
            }
        """)

        # Tab 1: ESA615 Device (if available)
        if ESA615_AVAILABLE:
            esa_tab = QWidget()
            esa_tab_layout = QVBoxLayout(esa_tab)
            esa_tab_layout.setContentsMargins(10, 10, 10, 10)
            esa_tab_layout.setSpacing(12)

            esa_label = QLabel("üîå ESA615 Device Files")
            esa_label.setStyleSheet("""
                background: transparent;
                font-weight: 700;
                font-size: 12pt;
                color: #10b981;
                padding-bottom: 8px;
            """)
            esa_tab_layout.addWidget(esa_label)

            # Create ESA615Widget with output folder and log callback
            self.esa615_widget = ESA615Widget(
                output_dir=os.path.abspath("."),
                log_callback=self.log_message
            )
            self.esa615_widget.files_downloaded.connect(self.add_downloaded_files)
            esa_tab_layout.addWidget(self.esa615_widget)

            tab_widget.addTab(esa_tab, "üîå ESA615 Device")

        # Tab 2: CSV Files
        csv_tab = QWidget()
        csv_tab_layout = QVBoxLayout(csv_tab)
        csv_tab_layout.setContentsMargins(10, 10, 10, 10)
        csv_tab_layout.setSpacing(12)

        file_group_label = QLabel("üìÅ CSV Files for Conversion")
        file_group_label.setStyleSheet("""
            background: transparent;
            font-weight: 700;
            font-size: 12pt;
            color: #3b82f6;
            padding-bottom: 8px;
        """)
        csv_tab_layout.addWidget(file_group_label)

        # File list with improved styling
        self.file_list = FileListWidget()
        self.file_list.setStyleSheet("""
            QListWidget {
                background: #f9fafb;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                padding: 8px;
                font-size: 10pt;
            }
            QListWidget::item {
                padding: 12px;
                border-radius: 6px;
                margin: 3px 0;
                background: white;
                border: 1px solid #e5e7eb;
            }
            QListWidget::item:hover {
                background: #f0fdf4;
                border-color: #10b981;
            }
            QListWidget::item:selected {
                background: #d1fae5;
                border-color: #10b981;
                color: #065f46;
            }
        """)
        csv_tab_layout.addWidget(self.file_list)

        # File buttons with icons
        file_btn_layout = QHBoxLayout()
        file_btn_layout.setSpacing(10)

        add_btn = QPushButton("üìÑ Add Files")
        add_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                padding: 10px 20px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563eb, stop:1 #1d4ed8);
            }
        """)
        add_btn.clicked.connect(self.add_files)
        file_btn_layout.addWidget(add_btn)

        remove_btn = QPushButton("üóë Remove")
        remove_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ef4444, stop:1 #dc2626);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #dc2626, stop:1 #b91c1c);
            }
        """)
        remove_btn.clicked.connect(self.remove_selected)
        file_btn_layout.addWidget(remove_btn)

        clear_btn = QPushButton("‚úñ Clear All")
        clear_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #6b7280, stop:1 #4b5563);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #4b5563, stop:1 #374151);
            }
        """)
        clear_btn.clicked.connect(self.file_list.clear)
        file_btn_layout.addWidget(clear_btn)

        csv_tab_layout.addLayout(file_btn_layout)
        tab_widget.addTab(csv_tab, "üìÅ CSV Files")

        layout.addWidget(tab_widget)

        # Configuration section in card (only output folder, template is fixed)
        config_card = QWidget()
        config_card.setStyleSheet("""
            QWidget {
                background: white;
                border-radius: 12px;
                border: 2px solid #e5e7eb;
            }
        """)
        config_layout = QVBoxLayout(config_card)
        config_layout.setContentsMargins(20, 15, 20, 15)
        config_layout.setSpacing(15)

        config_title = QLabel("‚öô Configuration")
        config_title.setStyleSheet("""
            font-weight: 700;
            font-size: 12pt;
            color: #7c3aed;
            padding-bottom: 8px;
        """)
        config_layout.addWidget(config_title)

        # Output folder
        output_label = QLabel("Output Folder:")
        output_label.setStyleSheet("font-weight: 600; color: #4b5563;")
        config_layout.addWidget(output_label)

        output_layout = QHBoxLayout()
        output_layout.setSpacing(10)
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("üìÇ Select output folder...")
        self.output_edit.setMinimumHeight(45)
        output_layout.addWidget(self.output_edit)

        output_btn = QPushButton("Browse")
        output_btn.setFixedWidth(100)
        output_btn.setMinimumHeight(45)
        output_btn.clicked.connect(self.browse_output)
        output_layout.addWidget(output_btn)
        config_layout.addLayout(output_layout)

        # Note about fixed template
        template_note = QLabel("üìã Template: example_Good.xlsx (fixed)")
        template_note.setStyleSheet("""
            color: #6b7280;
            font-size: 9pt;
            font-style: italic;
            padding: 8px;
            background: #f9fafb;
            border-radius: 6px;
        """)
        config_layout.addWidget(template_note)

        layout.addWidget(config_card)

        # Progress label in status bar style
        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet("""
            QLabel {
                background: #f3f4f6;
                border-radius: 8px;
                padding: 12px 15px;
                color: #374151;
                font-weight: 500;
                border: 1px solid #e5e7eb;
            }
        """)
        self.progress_label.setMinimumHeight(45)
        layout.addWidget(self.progress_label)

        # Convert button - Large and prominent
        self.convert_btn = QPushButton("üöÄ Convert Files to XLSX")
        self.convert_btn.setMinimumHeight(55)
        self.convert_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #10b981, stop:0.5 #059669, stop:1 #047857);
                color: white;
                font-size: 13pt;
                font-weight: 700;
                padding: 15px;
                border-radius: 12px;
                border: none;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #059669, stop:0.5 #047857, stop:1 #065f46);
            }
            QPushButton:pressed {
                background: #065f46;
            }
            QPushButton:disabled {
                background: #d1d5db;
                color: #9ca3af;
            }
        """)
        self.convert_btn.clicked.connect(self.start_conversion)
        layout.addWidget(self.convert_btn)

    def add_files(self):
        """Add CSV files via file dialog."""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Fluke CSV Files",
            "",
            "CSV Files (*.csv)"
        )
        for file in files:
            self.file_list.add_file(file)

    def add_downloaded_files(self, csv_files):
        """Add downloaded CSV files from ESA615 to file list."""
        for csv_path in csv_files:
            self.file_list.add_file(csv_path)
        self.log_message(f"‚úì Added {len(csv_files)} files to conversion list")

    def log_message(self, message):
        """Log message to progress label."""
        self.progress_label.setText(message)

    def remove_selected(self):
        """Remove selected files from list."""
        for item in self.file_list.selectedItems():
            self.file_list.takeItem(self.file_list.row(item))

    def browse_output(self):
        """Browse for output folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_edit.setText(folder)
            # Save to settings
            self.settings.setValue("output_folder", folder)
            # Update ESA615Widget output directory if available
            if ESA615_AVAILABLE and hasattr(self, 'esa615_widget'):
                self.esa615_widget.output_dir = folder

    def start_conversion(self):
        """Start conversion process."""
        # Validate inputs
        if self.file_list.count() == 0:
            QMessageBox.warning(self, "No Files", "Please add CSV files to convert.")
            return

        output_folder = self.output_edit.text().strip()
        if not output_folder or not os.path.exists(output_folder):
            QMessageBox.warning(self, "Invalid Output", "Please select a valid output folder.")
            return

        # Template is fixed to example_Good.xlsx
        template_path = self.template_path
        if not os.path.exists(template_path):
            QMessageBox.warning(self, "Template Missing",
                              f"Required template file not found: {template_path}\n"
                              f"Please ensure example_Good.xlsx is in the application directory.")
            return

        # Collect file paths
        csv_paths = []
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            csv_paths.append(item.data(Qt.UserRole))

        # Disable UI during conversion
        self.convert_btn.setEnabled(False)
        self.progress_label.setText("Starting conversion...")

        # Start conversion thread
        self.thread = ConvertThread(csv_paths, output_folder, template_path)
        self.thread.progress.connect(self.update_progress)
        self.thread.finished.connect(self.conversion_finished)
        self.thread.start()

    def update_progress(self, current: int, total: int, message: str):
        """Update progress label."""
        self.progress_label.setText(f"[{current}/{total}] {message}")

    def conversion_finished(self, success: int, errors: int, error_list: list):
        """Handle conversion completion."""
        self.convert_btn.setEnabled(True)
        self.progress_label.setText("")

        # Show summary
        msg = f"Conversion Complete\n\n"
        msg += f"Successful: {success}\n"
        msg += f"Errors: {errors}\n\n"

        if errors > 0:
            msg += "Error CSV generated with details.\n\n"
            msg += "First few errors:\n"
            for file, error in error_list[:5]:
                msg += f"- {file}: {error}\n"

        QMessageBox.information(self, "Conversion Complete", msg)


# ============================================================================
# SECTION 8: MAIN ENTRY POINT
# ============================================================================

def main():
    """Application entry point."""
    app = QApplication(sys.argv)
    app.setApplicationName("EST Converter")
    app.setOrganizationName("HugoNZ")

    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
